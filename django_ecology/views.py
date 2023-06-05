from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook

# Create your views here.
def home(request):
    if request.method == 'POST':
        return submit_proposal(request)
    else:
        return render(request, 'home.html')

def submit_proposal(request):
    title = request.POST.get("proposal_title")
    description = request.POST.get("proposal_description")

    try:
        workbook = load_workbook("ecology_proposals.xlsx")
    except FileNotFoundError:
        workbook = Workbook()

    worksheet = workbook.active

    if worksheet.max_row == 1:
        worksheet.append(["Title", "Description"])

    worksheet.append([title, description])

    workbook.save("ecology_proposals.xlsx")


    return HttpResponse("Proposal submitted successfully!")



